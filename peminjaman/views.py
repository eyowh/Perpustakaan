from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils.timezone import now
from .forms import *
from .models import Buku, Peminjaman, User
from django.db.models import Q
from django.contrib import messages
from django.utils import timezone
from datetime import datetime
from datetime import timedelta
from django.core.paginator import Paginator
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


def home_view(request):
    return render(request, 'home.html')


# Cek apakah user adalah admin
def is_admin(user):
    return user.is_staff or user.is_superuser

# --------------------- AUTH ---------------------

def login_view(request):
    if request.method == 'POST':
        form = LoginForm(data=request.POST)
        if form.is_valid():
            login(request, form.get_user())
            if is_admin(request.user):
                return redirect('dashboard_admin')
            else:
                return redirect('dashboard_user')
    else:
        form = LoginForm()
    return render(request, 'login.html', {'form': form})

def register_view(request):
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('dashboard_user')
    else:
        form = RegisterForm()
    return render(request, 'register.html', {'form': form})

def logout_view(request):
    logout(request)
    return redirect('home')

# --------------------- USER ---------------------

@login_required
def dashboard_user(request):
    buku_tersedia = Buku.objects.filter(stok_buku__gt=0)
    buku_populer = Buku.objects.order_by('-jumlah_dipinjam')[:5]

    hari_ini = timezone.now().date()
    peminjaman_user = Peminjaman.objects.filter(user=request.user, tanggal_dikembalikan__isnull=True)

    # Filter notifikasi
    jatuh_tempo_hari_ini = peminjaman_user.filter(tanggal_kembali=hari_ini)
    jatuh_tempo_besok = peminjaman_user.filter(tanggal_kembali=hari_ini + timedelta(days=1))
    sudah_terlambat = peminjaman_user.filter(tanggal_kembali__lt=hari_ini)

    return render(request, 'user/dashboarduser.html', {
        'nama': request.user.nama_panjang,
        'buku_tersedia': buku_tersedia,
        'buku_populer': buku_populer,
        'jatuh_tempo_hari_ini': jatuh_tempo_hari_ini,
        'jatuh_tempo_besok': jatuh_tempo_besok,
        'sudah_terlambat': sudah_terlambat,
        'buku_dipinjam': peminjaman_user,   # <-- tambah ini
    })

@login_required
def update_profil(request):
    user = request.user

    if request.method == 'POST':
        form = UpdateProfilForm(request.POST, instance=user)
        if form.is_valid():
            form.save()

            # Tandai bahwa user sudah pernah update profil
            if not user.sudah_update_profil:
                user.sudah_update_profil = True
                user.save(update_fields=['sudah_update_profil'])

            return redirect('update_profil')
    else:
        form = UpdateProfilForm(instance=user)

    return render(request, 'user/updateprofil.html', {'form': form})


@login_required
def pinjam_buku(request):
    user = request.user
    profil_lengkap = all([
        user.nama_panjang,
        user.jenis_kelamin,
        user.alamat,
        user.tgl_lahir
    ])

    query = request.GET.get('q', '')
    buku_list = Buku.objects.filter(
        Q(judul__icontains=query) |
        Q(penulis__icontains=query) |
        Q(penerbit__icontains=query),
        stok_buku__gt=0
    ) if query else Buku.objects.filter(stok_buku__gt=0)

    if request.method == 'POST':
        if not profil_lengkap:
            messages.error(request, "Silakan lengkapi profil Anda terlebih dahulu.")
            return redirect('update_profil')

        buku_id = request.POST.get('buku_id')
        tanggal_kembali = request.POST.get('tanggal_kembali')
        tanggal_pinjam = timezone.now().date()

        try:
            tgl_kembali = datetime.strptime(tanggal_kembali, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            messages.error(request, "Format tanggal tidak valid.")
            return redirect('pinjam_buku')

        if tgl_kembali <= tanggal_pinjam:
            messages.error(request, "Tanggal kembali harus setelah tanggal hari ini.")
            return redirect('pinjam_buku')

        buku = get_object_or_404(Buku, id=buku_id)

        # Buat peminjaman dengan status 'menunggu'
        Peminjaman.objects.create(
            user=request.user,
            buku=buku,
            tanggal_pinjam=tanggal_pinjam,
            tanggal_kembali=tgl_kembali,
            status='menunggu'  # <-- penting
        )

        messages.success(request, f'Permintaan peminjaman buku "{buku.judul}" telah diajukan dan menunggu persetujuan admin.')
        return redirect('pinjam_buku')

    return render(request, 'user/pinjambuku.html', {
        'buku_list': buku_list,
        'query': query,
        'profil_lengkap': profil_lengkap,
        'today': timezone.now().date().isoformat(),
        'nama': request.user.nama_panjang,
    })

@login_required
def kembalikan_buku(request, id):
    peminjaman = get_object_or_404(
        Peminjaman,
        id=id,
        user=request.user,
        tanggal_dikembalikan__isnull=True
    )

    peminjaman.tanggal_dikembalikan = timezone.now().date()
    peminjaman.save()

    # Kembalikan stok buku
    buku = peminjaman.buku
    buku.stok_buku += 1
    buku.save()

    messages.success(request, f'Buku "{buku.judul}" berhasil dikembalikan.')
    return redirect('dashboard_user')




@login_required
@login_required
def buku_dipinjam_user(request):
    today = timezone.localdate()
    denda_per_hari = 10000

    buku_dipinjam = Peminjaman.objects.filter(user=request.user).select_related('buku').order_by('-tanggal_pinjam')

    # Tambahkan properti denda sementara ke setiap objek
    for pem in buku_dipinjam:
        if pem.status == 'disetujui' and pem.tanggal_kembali < today and not pem.tanggal_dikembalikan:
            hari_telat = (today - pem.tanggal_kembali).days
            pem.denda_terhitung = hari_telat * denda_per_hari
        else:
            pem.denda_terhitung = 0

    return render(request, 'user/bukudipinjamuser.html', {
        'buku_dipinjam': buku_dipinjam,
        'nama': request.user.nama_panjang,
        'today': today,
    })

# --------------------- ADMIN ---------------------

@login_required
@user_passes_test(is_admin)
def dashboard_admin(request):
    total_buku = Buku.objects.count()
    total_pengguna = User.objects.filter(is_staff=False).count()
    peminjaman_aktif = Peminjaman.objects.filter(tanggal_dikembalikan__isnull=True).count()

    context = {
        'total_buku': total_buku,
        'total_pengguna': total_pengguna,
        'peminjaman_aktif': peminjaman_aktif,
    }
    return render(request, 'admin/dashboardadmin.html', context)

@login_required
@user_passes_test(is_admin)
def daftar_buku(request):
    buku_list = Buku.objects.all()
    return render(request, 'admin/daftarbuku.html', {'buku_list': buku_list})

@login_required
@user_passes_test(is_admin)
def tambah_buku(request):
    if request.method == 'POST':
        form = TambahBukuForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('daftar_buku')
    else:
        form = TambahBukuForm()
    return render(request, 'admin/tambahbuku.html', {'form': form})

@login_required
@user_passes_test(is_admin)
def edit_buku(request, id):
    buku = get_object_or_404(Buku, id=id)
    if request.method == 'POST':
        form = EditBukuForm(request.POST, request.FILES, instance=buku)
        if form.is_valid():
            form.save()
            return redirect('daftar_buku')
    else:
        form = EditBukuForm(instance=buku)
    return render(request, 'admin/editbuku.html', {'form': form})

@login_required
@user_passes_test(is_admin)
def hapus_buku(request, id):
    buku = get_object_or_404(Buku, id=id)
    buku.delete()
    return redirect('daftar_buku')

@login_required
@user_passes_test(is_admin)
def laporan_peminjaman(request):
    query = request.GET.get('q', '').strip()
    bulan = request.GET.get('bulan', '').strip()

    peminjaman_qs = Peminjaman.objects.all().order_by('-tanggal_pinjam')

    if query:
        peminjaman_qs = peminjaman_qs.filter(kode_laporan__icontains=query)

    if bulan:
        try:
            tahun, bulan_num = map(int, bulan.split('-'))
            peminjaman_qs = peminjaman_qs.filter(
                tanggal_pinjam__year=tahun,
                tanggal_pinjam__month=bulan_num
            )
        except ValueError:
            pass

    paginator = Paginator(peminjaman_qs, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'admin/laporanpeminjaman.html', {
        'peminjaman': page_obj,
        'query': query,
        'bulan': bulan,
    })


@login_required
@user_passes_test(is_admin)
def biodata_user(request):
    query = request.GET.get('q')
    users = User.objects.filter(is_staff=False)
    if query:
        users = users.filter(Q(username__icontains=query) | Q(nama_panjang__icontains=query))
    return render(request, 'admin/biodatauser.html', {'users': users})

def edit_user(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    if request.method == 'POST':
        user.username = request.POST.get('username')
        user.nama_panjang = request.POST.get('nama_panjang')
        user.email = request.POST.get('email')
        user.jenis_kelamin = request.POST.get('jenis_kelamin')
        user.tgl_lahir = request.POST.get('tgl_lahir')
        user.alamat = request.POST.get('alamat')
        user.save()
        return redirect('biodata_user')
    return render(request, 'admin/edit_user.html', {'user': user})

def hapus_user(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    user.delete()
    return redirect('biodata_user')

@login_required
@user_passes_test(is_admin)
def export_peminjaman_excel(request):
    # Ambil parameter bulan dari query string: format yyyy-mm
    bulan_param = request.GET.get("bulan")
    peminjamans = Peminjaman.objects.select_related("user", "buku")

    if bulan_param:
        try:
            # Convert ke bulan dan tahun
            tahun, bulan = map(int, bulan_param.split('-'))
            peminjamans = peminjamans.filter(tanggal_pinjam__year=tahun, tanggal_pinjam__month=bulan)
        except ValueError:
            pass  # Kalau format salah, tampilkan semua

    # Buat Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Peminjaman Bulan"

    headers = [
        "Kode Laporan", "Nama User", "Judul Buku",
        "Tanggal Pinjam", "Tanggal Kembali",
        "Tanggal Dikembalikan", "Status"
    ]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    for row_num, pinjam in enumerate(peminjamans, 2):
        ws.cell(row=row_num, column=1, value=pinjam.kode_laporan)
        ws.cell(row=row_num, column=2, value=pinjam.user.nama_panjang or pinjam.user.username)
        ws.cell(row=row_num, column=3, value=pinjam.buku.judul)
        ws.cell(row=row_num, column=4, value=str(pinjam.tanggal_pinjam))
        ws.cell(row=row_num, column=5, value=str(pinjam.tanggal_kembali))
        ws.cell(row=row_num, column=6, value=str(pinjam.tanggal_dikembalikan) if pinjam.tanggal_dikembalikan else "-")
        ws.cell(row=row_num, column=7, value=pinjam.status_jatuh_tempo())

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    nama_file = f"peminjaman_{bulan_param or 'semua'}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{nama_file}"'
    wb.save(response)
    return response



@login_required
@user_passes_test(is_admin)
def scan_pengembalian(request):
    return render(request, 'admin/scan_qr_pengembalian.html')

DENDA_PER_HARI = 10000

@login_required
@user_passes_test(is_admin)
def verifikasi_pengembalian(request, kode):
    peminjaman = get_object_or_404(
        Peminjaman,
        kode_laporan=kode,
        status='disetujui',
        tanggal_dikembalikan__isnull=True
    )

    today = timezone.now().date()
    hari_telat = 0
    perkiraan_denda = 0

    # Hitung denda hanya jika terlambat (GET dan POST sama)
    if today > peminjaman.tanggal_kembali:
        hari_telat = (today - peminjaman.tanggal_kembali).days
        perkiraan_denda = hari_telat * DENDA_PER_HARI

    if request.method == 'POST':
        peminjaman.tanggal_dikembalikan = today
        peminjaman.status = 'dikembalikan'
        peminjaman.denda = perkiraan_denda

        peminjaman.save()

        # Tambahkan stok buku kembali
        buku = peminjaman.buku
        buku.stok_buku += 1
        buku.save()

        # Feedback ke admin
        if perkiraan_denda > 0:
            messages.warning(
                request,
                f"Buku dikembalikan dengan denda Rp{perkiraan_denda:,} untuk keterlambatan {hari_telat} hari."
            )
        else:
            messages.success(request, "Buku berhasil dikembalikan tanpa denda.")

        return redirect('dashboard_admin')

    return render(request, 'admin/verifikasi_pengembalian.html', {
        'peminjaman': peminjaman,
        'today': today,
        'hari_telat': hari_telat,
        'perkiraan_denda': perkiraan_denda,
    })

@login_required
@user_passes_test(is_admin)
def scan_peminjaman(request):
    return render(request, 'admin/scan_qr_peminjaman.html')

@login_required
@user_passes_test(is_admin)
def verifikasi_peminjaman(request):
    kode = request.GET.get('kode')
    try:
        peminjaman = Peminjaman.objects.get(kode_laporan=kode, status='menunggu')
    except Peminjaman.DoesNotExist:
        messages.error(request, "Peminjaman tidak ditemukan atau sudah diproses.")
        return redirect('scan_peminjaman')

    if request.method == 'POST':
        aksi = request.POST.get('aksi')
        if aksi == 'setujui':
            peminjaman.status = 'disetujui'
            peminjaman.buku.stok_buku -= 1
            peminjaman.buku.jumlah_dipinjam += 1
            peminjaman.buku.save()
            peminjaman.save()
            messages.success(request, "Peminjaman disetujui.")
        elif aksi == 'tolak':
            peminjaman.status = 'ditolak'
            peminjaman.save()
            messages.warning(request, "Peminjaman ditolak.")
        return redirect('dashboard_admin')

    return render(request, 'admin/verifikasi_peminjaman.html', {
        'peminjaman': peminjaman
    })

@login_required
@user_passes_test(is_admin)
def konfirmasi_denda(request, id):
    peminjaman = get_object_or_404(Peminjaman, id=id)

    if request.method == 'POST':
        form = KonfirmasiDendaForm(request.POST, instance=peminjaman)
        if form.is_valid():
            form.save()
            messages.success(request, "Status denda berhasil diperbarui.")
            return redirect('laporan_peminjaman')
    else:
        form = KonfirmasiDendaForm(instance=peminjaman)

    return render(request, 'admin/konfirmasi_denda.html', {
        'peminjaman': peminjaman,
        'form': form
    })